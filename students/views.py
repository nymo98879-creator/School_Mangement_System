from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.contrib.auth import get_user_model
from django.core.mail import send_mail
from django.conf import settings
from django.urls import reverse  # ADD THIS IMPORT
from accounts.decorators import admin_required
from .models import Student
from .forms import StudentForm, StudentSearchForm, StudentBulkUploadForm, StudentSelfRegistrationForm
from classes.models import Class
import csv
import openpyxl
from datetime import datetime
import json
import qrcode
from io import BytesIO
from django.core.files.base import ContentFile

User = get_user_model()

# ... rest of your code (all functions remain the same) ...


# ============ DASHBOARD ============

@login_required
@admin_required
def student_dashboard(request):
    """Admin dashboard with student management"""
    context = {
        'total_students': Student.objects.count(),
        'active_students': Student.objects.filter(is_active=True).count(),
        'inactive_students': Student.objects.filter(is_active=False).count(),
        'qr_registered': Student.objects.filter(qr_registered=True).count(),
        'qr_not_registered': Student.objects.filter(qr_registered=False).count(),
        'recent_students': Student.objects.all().order_by('-created_at')[:5],
    }
    return render(request, 'Backend/admin/student/student_dashboard.html', context)

# ============ STUDENT LIST ============

@login_required
@admin_required
def student_list(request):
    """Display list of all students with search and filter"""
    students = Student.objects.all()
    form = StudentSearchForm(request.GET or None)
    
    if form.is_valid():
        search = form.cleaned_data.get('search')
        class_filter = form.cleaned_data.get('class_filter')
        gender_filter = form.cleaned_data.get('gender_filter')
        status_filter = form.cleaned_data.get('status_filter')
        qr_status = form.cleaned_data.get('qr_status')
        
        if search:
            students = students.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(student_id__icontains=search) |
                Q(email__icontains=search)
            )
        
        if class_filter:
            students = students.filter(class_enrolled=class_filter)
        
        if gender_filter:
            students = students.filter(gender=gender_filter)
        
        if status_filter == 'active':
            students = students.filter(is_active=True)
        elif status_filter == 'inactive':
            students = students.filter(is_active=False)
        
        if qr_status == 'registered':
            students = students.filter(qr_registered=True)
        elif qr_status == 'not_registered':
            students = students.filter(qr_registered=False)
    
    paginator = Paginator(students, 10)
    page_number = request.GET.get('page')
    students_page = paginator.get_page(page_number)
    
    context = {
        'students': students_page,
        'form': form,
        'total_students': Student.objects.count(),
        'active_students': Student.objects.filter(is_active=True).count(),
        'inactive_students': Student.objects.filter(is_active=False).count(),
        'qr_not_registered': Student.objects.filter(qr_registered=False).count(),
        'qr_registered_count': Student.objects.filter(qr_registered=True).count(),
    }
    return render(request, 'Backend/admin/student/student_list.html', context)

# ============ ADD STUDENT ============

@login_required
@admin_required
def student_add(request):
    """Add a new student"""
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES, is_edit=False)
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data.get('password')
            first_name = form.cleaned_data['first_name']
            last_name = form.cleaned_data['last_name']
            
            try:
                user = User.objects.create_user(
                    username=email,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name
                )
            except Exception as e:
                messages.error(request, f'Error creating user account: {str(e)}')
                context = {
                    'form': form,
                    'title': 'Add New Student',
                    'is_edit': False,
                }
                return render(request, 'Backend/admin/student/student_form.html', context)
            
            student = form.save(commit=False)
            student.user = user
            student.save()
            
            messages.success(request, f'✅ Student {student.get_full_name()} added successfully! Student ID: {student.student_id}')
            return redirect('students:student_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = StudentForm(is_edit=False)
    
    context = {
        'form': form,
        'title': 'Add New Student',
        'is_edit': False,
    }
    return render(request, 'Backend/admin/student/student_form.html', context)

# ============ BULK UPLOAD ============

@login_required
@admin_required
def student_bulk_upload(request):
    """Bulk upload students from Excel/CSV"""
    if request.method == 'POST':
        form = StudentBulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            default_class = form.cleaned_data['class_enrolled']
            
            if excel_file.name.endswith('.csv'):
                data = csv.DictReader(excel_file.read().decode('utf-8').splitlines())
                students_data = list(data)
            else:
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active
                headers = [cell.value for cell in sheet[1]]
                students_data = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    student_dict = dict(zip(headers, row))
                    students_data.append(student_dict)
            
            success_count = 0
            error_count = 0
            errors = []
            
            for student_data in students_data:
                try:
                    first_name = student_data.get('first_name', '').strip()
                    last_name = student_data.get('last_name', '').strip()
                    email = student_data.get('email', '').strip().lower()
                    phone = student_data.get('phone', '').strip()
                    date_of_birth = student_data.get('date_of_birth', '')
                    gender = student_data.get('gender', '').strip().upper()
                    guardian_name = student_data.get('guardian_name', '').strip()
                    guardian_phone = student_data.get('guardian_phone', '').strip()
                    guardian_email = student_data.get('guardian_email', '').strip()
                    address = student_data.get('address', '').strip()
                    
                    if not all([first_name, last_name, email, phone, date_of_birth, gender, guardian_name, guardian_phone]):
                        error_count += 1
                        errors.append(f"Missing required fields for {first_name} {last_name}")
                        continue
                    
                    if gender not in ['M', 'F', 'O']:
                        error_count += 1
                        errors.append(f"Invalid gender '{gender}' for {first_name} {last_name}")
                        continue
                    
                    if Student.objects.filter(email=email).exists():
                        error_count += 1
                        errors.append(f"Email {email} already exists")
                        continue
                    
                    password = User.objects.make_random_password()
                    user = User.objects.create_user(
                        username=email,
                        email=email,
                        password=password,
                        first_name=first_name,
                        last_name=last_name
                    )
                    
                    student = Student(
                        user=user,
                        first_name=first_name,
                        last_name=last_name,
                        email=email,
                        phone=phone,
                        date_of_birth=date_of_birth,
                        gender=gender,
                        guardian_name=guardian_name,
                        guardian_phone=guardian_phone,
                        guardian_email=guardian_email if guardian_email else None,
                        address=address if address else '',
                        class_enrolled=default_class
                    )
                    student.save()
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f"Error processing {student_data.get('first_name', 'Unknown')}: {str(e)}")
            
            message = f'✅ Successfully added {success_count} students.'
            if error_count > 0:
                message += f' ❌ Failed: {error_count}.'
                messages.warning(request, message)
                if errors:
                    messages.error(request, f'Errors: {", ".join(errors[:5])}...')
            else:
                messages.success(request, message)
            
            return redirect('students:student_list')
    else:
        form = StudentBulkUploadForm()
    
    context = {
        'form': form,
        'title': 'Bulk Upload Students',
    }
    return render(request, 'Backend/admin/student/student_bulk_upload.html', context)

# ============ STUDENT DETAIL ============

@login_required
def student_detail(request, pk):
    """View student details with QR code"""
    student = get_object_or_404(Student, pk=pk)
    
    # Check permission
    if request.user.role != 'admin' and request.user != student.user:
        messages.error(request, 'You do not have permission to view this student.')
        return redirect('dashboard')
    
    if not student.qr_code:
        student.generate_qr_code()
        student.save(update_fields=['qr_code'])
    
    context = {
        'student': student,
    }
    return render(request, 'Backend/admin/student/student_detail.html', context)

# ============ DOWNLOAD QR CODE ============

@login_required
def download_qr_code(request, pk):
    """Download QR code for a specific student"""
    student = get_object_or_404(Student, pk=pk)
    
    # Check permission
    if request.user.role != 'admin' and request.user != student.user:
        messages.error(request, 'You do not have permission to download this QR code.')
        return redirect('students:student_detail', pk=pk)
    
    if not student.qr_code:
        student.generate_qr_code()
        student.save(update_fields=['qr_code'])
    
    response = HttpResponse(student.qr_code.read(), content_type='image/png')
    response['Content-Disposition'] = f'attachment; filename="{student.student_id}_qr.png"'
    return response

# ============ GENERATE QR FOR ALL STUDENTS ============

@login_required
@admin_required
def generate_all_qr(request):
    """Generate QR codes for ALL students at once"""
    if request.method == 'POST':
        students = Student.objects.all()
        count = 0
        
        for student in students:
            student.generate_qr_code()
            student.save(update_fields=['qr_code'])
            count += 1
        
        if count > 0:
            messages.success(request, f'✅ QR codes generated for ALL {count} students!')
        else:
            messages.info(request, 'ℹ️ No students found to generate QR codes.')
        
        return redirect('students:student_list')
    
    # GET request - Show confirmation page
    total_students = Student.objects.count()
    context = {
        'total_students': total_students,
    }
    return render(request, 'Backend/admin/student/generate_all_qr_confirm.html', context)

# ============ STUDENT EDIT ============

@login_required
@admin_required
def student_edit(request, pk):
    """Edit an existing student"""
    student = get_object_or_404(Student, pk=pk)
    
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES, instance=student, is_edit=True)
        if form.is_valid():
            updated_student = form.save()
            
            if updated_student.user:
                user = updated_student.user
                user.first_name = form.cleaned_data['first_name']
                user.last_name = form.cleaned_data['last_name']
                user.email = form.cleaned_data['email']
                user.username = form.cleaned_data['email']
                
                password = form.cleaned_data.get('password')
                if password:
                    user.set_password(password)
                
                user.save()
            
            messages.success(request, f'✅ Student {updated_student.get_full_name()} updated successfully!')
            return redirect('students:student_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = StudentForm(instance=student, is_edit=True)
    
    context = {
        'form': form,
        'title': 'Edit Student',
        'student': student,
        'is_edit': True,
    }
    return render(request, 'Backend/admin/student/student_form.html', context)

# ============ STUDENT DELETE ============

@login_required
@admin_required
def student_delete(request, pk):
    """Delete a student"""
    student = get_object_or_404(Student, pk=pk)
    
    if request.method == 'POST':
        student_name = student.get_full_name()
        if student.user:
            student.user.delete()
        student.delete()
        messages.success(request, f'🗑️ Student {student_name} deleted successfully!')
        return redirect('students:student_list')
    
    context = {
        'student': student,
    }
    return render(request, 'Backend/admin/student/student_delete.html', context)

# ============ STUDENT TOGGLE STATUS ============

@login_required
@admin_required
def student_toggle_status(request, pk):
    """Toggle student active status"""
    student = get_object_or_404(Student, pk=pk)
    student.is_active = not student.is_active
    student.save()
    
    if student.user:
        student.user.is_active = student.is_active
        student.user.save()
    
    status = 'activated' if student.is_active else 'deactivated'
    messages.success(request, f'Student {student.get_full_name()} {status} successfully!')
    return redirect('students:student_list')

# ============ STUDENT EXPORT CSV ============

@login_required
@admin_required
def student_export_csv(request):
    """Export student data to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="students_{datetime.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'ID', 'Student ID', 'Full Name', 'Email', 'Phone', 'Gender', 
        'Date of Birth', 'Class', 'Guardian Name', 'Guardian Phone',
        'Guardian Email', 'Address', 'QR Registered', 'Enrollment Date', 'Status'
    ])
    
    students = Student.objects.all()
    for student in students:
        writer.writerow([
            student.id,
            student.student_id,
            student.get_full_name(),
            student.email,
            student.phone,
            student.get_gender_display(),
            student.date_of_birth,
            student.class_enrolled.name if student.class_enrolled else 'Not Assigned',
            student.guardian_name,
            student.guardian_phone,
            student.guardian_email or '',
            student.address,
            'Yes' if student.qr_registered else 'No',
            student.enrollment_date,
            'Active' if student.is_active else 'Inactive'
        ])
    
    return response

# ============ DOWNLOAD TEMPLATE ============

@login_required
@admin_required
def download_template(request):
    """Download Excel template for bulk upload"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Students'
    
    headers = [
        'first_name', 'last_name', 'email', 'phone', 'date_of_birth',
        'gender', 'guardian_name', 'guardian_phone', 'guardian_email', 'address'
    ]
    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col, value=header)
        sheet.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
    
    example = [
        'John', 'Doe', 'john.doe@example.com', '+1234567890', '2000-01-01',
        'M', 'Jane Doe', '+1234567891', 'jane.doe@example.com', '123 Main St'
    ]
    for col, value in enumerate(example, 1):
        sheet.cell(row=2, column=col, value=value)
    
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(response)
    return response

# ============ STUDENT SELF-REGISTRATION ============

def student_self_register_form(request):
    """Public page for students to register themselves"""
    if request.method == 'POST':
        form = StudentSelfRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            first_name = form.cleaned_data['first_name']
            last_name = form.cleaned_data['last_name']
            
            try:
                user = User.objects.create_user(
                    username=email,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name
                )
            except Exception as e:
                messages.error(request, f'Error creating account: {str(e)}')
                context = {'form': form}
                return render(request, 'Frontend/student/self_register.html', context)
            
            student = form.save(commit=False)
            student.user = user
            student.qr_registered = True
            student.registration_completed = True
            student.is_active = True
            student.save()
            
            from django.contrib.auth import login as auth_login
            auth_login(request, user)
            
            messages.success(request, f'✅ Welcome {student.get_full_name()}! Your registration is complete.')
            return redirect('students:student_register_success')
        else:
            messages.error(request, 'Please correct the errors below.')
            context = {'form': form}
            return render(request, 'Frontend/student/self_register.html', context)
    else:
        form = StudentSelfRegistrationForm()
    
    context = {'form': form}
    return render(request, 'Frontend/student/self_register.html', context)

def student_register_success(request):
    """Success page after self-registration"""
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        student = None
    
    context = {
        'student': student,
    }
    return render(request, 'Frontend/student/register_success.html', context)

# ============ VERIFY STUDENT QR ============

def verify_student_qr(request):
    """API endpoint to verify student QR code"""
    if request.method == 'POST':
        qr_data = request.POST.get('qr_data', '')
        try:
            data = json.loads(qr_data)
            student_id = data.get('student_id')
            
            if not student_id:
                return JsonResponse({'success': False, 'error': 'Invalid QR code'})
            
            student = get_object_or_404(Student, student_id=student_id)
            
            return JsonResponse({
                'success': True,
                'student': {
                    'id': student.id,
                    'student_id': student.student_id,
                    'name': student.get_full_name(),
                    'email': student.email,
                    'class': student.class_enrolled.name if student.class_enrolled else 'Not Assigned',
                    'is_active': student.is_active,
                    'profile_picture': student.profile_picture.url if student.profile_picture else None,
                }
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request'})

# ============ STUDENT SCAN REGISTER ============

@login_required
def student_scan_register(request):
    """Page for student to scan QR code and register"""
    return render(request, 'Frontend/student/scan_register.html')

@login_required
def student_self_register(request, student_id):
    """Student scans QR code and completes registration"""
    student = get_object_or_404(Student, student_id=student_id)
    
    if request.user != student.user:
        messages.error(request, 'You are not authorized to register this student.')
        return redirect('home')
    
    if request.method == 'POST':
        student.qr_registered = True
        student.save()
        messages.success(request, f'✅ Registration completed for {student.get_full_name()}!')
        return redirect('students:student_dashboard')
    
    context = {'student': student}
    return render(request, 'Backend/admin/student/student_self_register.html', context)


# Add these functions to students/views.py
# ============ GENERATE REGISTRATION QR ============

@login_required
@admin_required
def generate_registration_qr(request):
    """Generate QR code for student self-registration"""
    if request.method == 'POST':
        # Get the registration URL using reverse
        registration_url = request.build_absolute_uri(reverse('students:student_self_register_form'))
        
        # Create QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        
        qr_data = {
            'type': 'student_registration',
            'url': registration_url,
            'message': 'Scan to register as a student'
        }
        qr.add_data(json.dumps(qr_data))
        qr.make(fit=True)
        
        qr_image = qr.make_image(fill_color="black", back_color="white")
        
        buffer = BytesIO()
        qr_image.save(buffer, 'PNG')
        
        # Save to session for download
        request.session['registration_qr'] = buffer.getvalue().hex()
        
        # Convert to base64 for display
        import base64
        qr_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
        
        context = {
            'qr_generated': True,
            'registration_url': registration_url,
            'qr_code_base64': qr_base64,
        }
        return render(request, 'Backend/admin/student/generate_registration_qr.html', context)
    
    context = {
        'qr_generated': False,
    }
    return render(request, 'Backend/admin/student/generate_registration_qr.html', context)


@login_required
@admin_required
def download_registration_qr(request):
    """Download the registration QR code"""
    qr_data = request.session.get('registration_qr')
    
    if not qr_data:
        # Generate new QR if not in session
        registration_url = request.build_absolute_uri(reverse('students:student_self_register_form'))
        
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        
        qr_data_dict = {
            'type': 'student_registration',
            'url': registration_url,
            'message': 'Scan to register as a student'
        }
        qr.add_data(json.dumps(qr_data_dict))
        qr.make(fit=True)
        
        qr_image = qr.make_image(fill_color="black", back_color="white")
        
        buffer = BytesIO()
        qr_image.save(buffer, 'PNG')
        qr_bytes = buffer.getvalue()
    else:
        qr_bytes = bytes.fromhex(qr_data)
    
    response = HttpResponse(qr_bytes, content_type='image/png')
    response['Content-Disposition'] = 'attachment; filename="student_registration_qr.png"'
    return response